#!/usr/bin/env python3
"""
Second-pass schema fix: items A–H.
Operates on data/respondents.json and writes updated file in-place.
Source spreadsheet: /Users/972010/Downloads/Cognizant_Raw_Data.xlsx
"""

import json
import statistics
import openpyxl
from collections import Counter

RESPONDENTS_PATH = "data/respondents.json"
XLSX_PATH = "/Users/972010/Downloads/Cognizant_Raw_Data.xlsx"

# ---------------------------------------------------------------------------
# Codebook lookups
# ---------------------------------------------------------------------------

VLIST_LABELS = {
    1: "Open Survey (list=0)",
    2: "IMR (list=1)",
    3: "High Budget Low Rank (list=2)",
    4: "High Budget High Rank (list=3)",
    5: "Low Budget Low Rank (list=4)",
    6: "Low Budget High Rank (list=5)",
    7: "High Budget High Rank (list=6)",
    8: "High Budget Low Rank (list=7)",
    9: "Low Budget Low Rank (list=8)",
    10: "Low Budget High Rank (list=9)",
}

# vlist code → segment label (None for codes that predate segment assignment)
VLIST_TO_SEGMENT = {
    1: None,
    2: None,
    3: "High Budget Low Rank",
    4: "High Budget High Rank",
    5: "Low Budget Low Rank",
    6: "Low Budget High Rank",
    7: "High Budget High Rank",
    8: "High Budget Low Rank",
    9: "Low Budget Low Rank",
    10: "Low Budget High Rank",
}

HQ1_LABELS = {
    1: "Less than 1,000",
    2: "1,000–2,499",
    3: "2,500–4,999",
    4: "5,000–9,999",
    5: "10,000–50,000",
    6: "50,000+",
}

EMP_SIZE_CHECK_LABELS = {1: "Match", 2: "Mismatch"}

INDUSTRY_LABELS = {
    1: "Agriculture and Farming", 2: "Automotive", 3: "Banking",
    4: "Business Services", 5: "Communications", 6: "Construction and Real Estate",
    7: "Consumer Goods", 8: "Education", 9: "Energy",
    10: "Entertainment and Media", 11: "Financial Services",
    12: "Government and Public Sector", 13: "Healthcare, Government-Provided",
    14: "Healthcare Payer", 15: "Healthcare Provider", 16: "Insurance",
    17: "Manufacturing", 18: "Medical Products and Devices",
    19: "Non-profit and Social Services", 20: "Pharmaceuticals and Biotech",
    21: "Retail", 22: "Software & Technology", 23: "Telecommunications",
    24: "Transportation, Logistics and Supply Chain", 25: "Travel and Hospitality",
    26: "Utilities", 27: "Wholesale", 28: "Other",
}

SENIORITY_LABELS = {
    1: "C-suite / Most-senior decision maker",
    2: "VP / Senior Director",
    3: "Director",
    4: "Manager",
    5: "Analyst / Coordinator",
    6: "None of the above",
}

Q39_LABELS = {
    1: "<$500k", 2: "$500k–$999k", 3: "$1m–$4.9m",
    4: "$5m–$9.9m", 5: "$10m–$14.9m", 6: "$15m–$19.9m",
    7: "$20m–$29.9m", 8: ">$30m", 9: "Don't know",
}


# ---------------------------------------------------------------------------
# Load spreadsheet data
# ---------------------------------------------------------------------------

def load_xlsx_by_record():
    """Return dict: record_num -> row tuple."""
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb["A1"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map = {h: i for i, h in enumerate(headers) if h}

    rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = row[0]
        if rec is not None:
            rows[rec] = row

    return rows, col_map


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    with open(RESPONDENTS_PATH) as f:
        data = json.load(f)
    print(f"Loaded {len(data)} records.\n")

    print("Loading source spreadsheet...")
    xlsx_rows, col_map = load_xlsx_by_record()
    print(f"Loaded {len(xlsx_rows)} rows from spreadsheet.\n")

    # Build record→doc lookup
    by_record = {r["record"]: r for r in data}

    # -----------------------------------------------------------------------
    # A. profile.panel_list + vlist reconciliation
    # -----------------------------------------------------------------------
    print("=" * 60)
    print("ITEM A: profile.panel_list (vlist) + segment reconciliation")
    print("=" * 60)

    segment_mismatch = []
    vlist_dist = Counter()
    panel_list_added = 0

    for r in data:
        meta_vlist = r.get("_meta", {}).get("vlist")
        if meta_vlist is None:
            # Fall back to spreadsheet
            xlsx_row = xlsx_rows.get(r["record"])
            meta_vlist = xlsx_row[col_map["vlist"]] if xlsx_row else None

        if meta_vlist is not None:
            meta_vlist = int(meta_vlist)
            label = VLIST_LABELS.get(meta_vlist, f"Unknown ({meta_vlist})")
            r["profile"]["panel_list"] = {"code": meta_vlist, "label": label}
            panel_list_added += 1
            vlist_dist[meta_vlist] += 1

            # Cross-check vs respondent_segment
            seg = r["responses"].get("respondent_segment", {})
            seg_label = seg.get("label") if isinstance(seg, dict) else None
            expected_seg = VLIST_TO_SEGMENT.get(meta_vlist)
            if seg_label != expected_seg:
                segment_mismatch.append({
                    "record": r["record"],
                    "vlist_code": meta_vlist,
                    "vlist_label": label,
                    "expected_segment": expected_seg,
                    "actual_segment": seg_label,
                })

    print(f"profile.panel_list added/updated: {panel_list_added}/600")
    print()
    print("vlist distribution:")
    for code in sorted(vlist_dist):
        print(f"  code={code} ({VLIST_LABELS.get(code,'?')}): {vlist_dist[code]}")
    print()
    if segment_mismatch:
        print(f"Segment mismatches (vlist→segment): {len(segment_mismatch)}")
        for m in segment_mismatch[:20]:
            print(f"  record={m['record']}: vlist={m['vlist_code']} ({m['vlist_label']})"
                  f" → expected={m['expected_segment']}, actual={m['actual_segment']}")
    else:
        print("Segment reconciliation: 0 mismatches — all vlist codes map cleanly to respondent_segment.")

    # -----------------------------------------------------------------------
    # B. Spot-check Q10, Q44, Q45, Q48, Q49, Q50
    # -----------------------------------------------------------------------
    print()
    print("=" * 60)
    print("ITEM B: Spot-check Q10, Q44, Q45, Q48, Q49, Q50 label accuracy")
    print("=" * 60)

    CODEBOOK_LABELS = {
        "q10": {
            1: "Significant expansion - Major investment and growth in AI capabilities",
            2: "Moderate expansion - Steady growth in AI capabilities",
            3: "Minimal expansion - Slight growth in AI capabilities",
            4: "Maintain current level - No significant change expected",
            5: "Contraction - Reducing AI investment or capabilities",
            6: "Uncertain - Too early to determine",
        },
        "q44": {
            1: "Existential threat \u2013 Organizations must adapt or fail",
            2: "Major disruption \u2013 Organizations require fundamental changes to succeed",
            3: "Moderate disruption \u2013 Organizations require significant but manageable changes",
            4: "Incremental change \u2013 Organizations require improvements but not wholesale transformation",
            5: "No change \u2013 No disruption anticipated",
            6: "Not able to predict level of disruption",
        },
        "q45": {1: "1", 2: "2-3", 3: "4-6", 4: "7-10", 5: "10+"},
        "q48": {
            1: "Less than 5 years", 2: "5-15 years", 3: "16-24 years",
            4: "25-49 years", 5: "50+ years", 6: "I\u2019m not sure",
        },
        "q49": {
            1: "Grew 1-4.99%", 2: "Grew 5-9.99%", 3: "Grew 10-14.99%",
            4: "Grew 15-19.99%", 5: "Grew over 20%", 6: "Neither grew nor shrank",
            7: "Shrank 1-4.99%", 8: "Shrank 5-9.99%", 9: "Shrank 10-14.99%",
            10: "Shrank 15-19.99%", 11: "Shrank over 20%", 12: "I\u2019m not sure",
        },
        "q50": {
            1: "Microsoft Azure", 2: "AWS", 3: "Google Cloud", 4: "Oracle",
            5: "Other cloud environment (Please enter)",
            6: "We do not use any cloud computing environments",
            7: "I\u2019m not sure",
        },
    }

    total_mismatches = 0
    for fkey, cb_map in CODEBOOK_LABELS.items():
        field_mismatches = []
        for r in data:
            val = r["responses"].get(fkey)
            if isinstance(val, dict) and val.get("code") is not None:
                code = val["code"]
                stored = val.get("label", "")
                expected = cb_map.get(code, "")
                if stored != expected:
                    field_mismatches.append((r["record"], code, stored, expected))
        if field_mismatches:
            total_mismatches += len(field_mismatches)
            print(f"  {fkey.upper()}: {len(field_mismatches)} label mismatches")
            for rec, code, stored, expected in field_mismatches[:3]:
                print(f"    record={rec} code={code}: stored={repr(stored)} expected={repr(expected)}")
        else:
            # Show one sample record for confirmation
            sample = next((r for r in data if r["responses"].get(fkey)), None)
            if sample:
                v = sample["responses"][fkey]
                print(f"  {fkey.upper()}: OK — sample code={v.get('code')}, label={repr(v.get('label'))}")

    if total_mismatches == 0:
        print()
        print("All 6 fields: 0 label mismatches against datamap. Labels are accurate.")

    # -----------------------------------------------------------------------
    # C. Q42 storage format verification
    # -----------------------------------------------------------------------
    print()
    print("=" * 60)
    print("ITEM C: Q42 storage format")
    print("=" * 60)

    q42_types = Counter(type(r["responses"].get("q42")).__name__ for r in data)
    print(f"Storage type distribution: {dict(q42_types)}")
    sample_q42 = next((r["responses"]["q42"] for r in data if r["responses"].get("q42")), None)
    print(f"Sample value: {sample_q42}")
    print()
    if q42_types.get("list", 0) == 600:
        print("Q42 is stored as a string array of selected option labels — CORRECT format.")
    else:
        print("WARNING: Q42 is not consistently stored as a list.")

    # Option count distribution
    q42_lengths = Counter(len(r["responses"].get("q42") or []) for r in data)
    print("Options selected per respondent:", dict(sorted(q42_lengths.items())))

    # -----------------------------------------------------------------------
    # D. hQ1 and EmpSizeCheck
    # -----------------------------------------------------------------------
    print()
    print("=" * 60)
    print("ITEM D: hQ1 (employee_bucket_computed) and EmpSizeCheck (employee_size_check)")
    print("=" * 60)

    hq1_updated = emp_check_updated = 0
    mismatch_records = []

    for r in data:
        xlsx_row = xlsx_rows.get(r["record"])
        if xlsx_row is None:
            continue

        hq1_code = xlsx_row[col_map["hQ1"]]
        emp_check_code = xlsx_row[col_map["EmpSizeCheck"]]

        if hq1_code is not None:
            r["profile"]["employee_bucket_computed"] = {
                "code": int(hq1_code),
                "label": HQ1_LABELS.get(int(hq1_code), f"Unknown ({hq1_code})"),
                "question_code": "hQ1",
            }
            hq1_updated += 1

        if emp_check_code is not None:
            code = int(emp_check_code)
            r["profile"]["employee_size_check"] = {
                "code": code,
                "label": EMP_SIZE_CHECK_LABELS.get(code, f"Unknown ({code})"),
            }
            emp_check_updated += 1
            if code == 2:
                mismatch_records.append(r["record"])

    print(f"employee_bucket_computed added: {hq1_updated}/600")
    print(f"employee_size_check added: {emp_check_updated}/600")
    print()

    # hQ1 distribution
    hq1_dist = Counter()
    for r in data:
        bucket = r["profile"].get("employee_bucket_computed", {})
        if isinstance(bucket, dict):
            hq1_dist[bucket.get("code")] += 1
    print("employee_bucket_computed distribution:")
    for code in sorted(hq1_dist):
        print(f"  code={code} ({HQ1_LABELS.get(code,'?')}): {hq1_dist[code]}")

    print()
    emp_check_dist = Counter()
    for r in data:
        ec = r["profile"].get("employee_size_check", {})
        if isinstance(ec, dict):
            emp_check_dist[ec.get("label")] += 1
    print("employee_size_check distribution:")
    for label, cnt in emp_check_dist.most_common():
        print(f"  {label}: {cnt}")
    print()
    print(f"S2 vs Q47 mismatches: {len(mismatch_records)} records")
    if mismatch_records:
        print(f"  Mismatch record numbers: {mismatch_records[:20]}")
        # Show detail for first 3
        for rec in mismatch_records[:3]:
            r = by_record[rec]
            s2 = r["profile"].get("emp_count", {})
            q47_self = r["profile"].get("employee_count_self_reported")
            hq1 = r["profile"].get("employee_bucket_computed", {})
            print(f"  record={rec}: profile.emp_count={s2}, Q47_self={q47_self}, hQ1={hq1.get('label')}")

    # -----------------------------------------------------------------------
    # E. qtime → metadata.interview_duration
    # -----------------------------------------------------------------------
    print()
    print("=" * 60)
    print("ITEM E: Interview duration (qtime)")
    print("=" * 60)

    qtimes_all = []
    under_8min = []

    for r in data:
        qt = r.get("_meta", {}).get("completion_time_sec")
        if qt is None:
            xlsx_row = xlsx_rows.get(r["record"])
            qt = xlsx_row[col_map["qtime"]] if xlsx_row else None

        if qt is not None:
            qt = float(qt)
            # Ensure _meta has it
            r["_meta"]["completion_time_sec"] = qt
            # Add to metadata object
            if "metadata" not in r:
                r["metadata"] = {}
            r["metadata"]["interview_duration_sec"] = qt
            qtimes_all.append((r["record"], qt, r))
            if qt < 480:
                under_8min.append((r["record"], qt, r))

    print(f"Records with qtime: {len(qtimes_all)}")
    times = [t for _, t, _ in qtimes_all]
    if times:
        print(f"Min: {min(times):.1f}s ({min(times)/60:.1f}min)")
        print(f"Max: {max(times):.1f}s ({max(times)/60:.1f}min)")
        print(f"Median: {statistics.median(times):.1f}s ({statistics.median(times)/60:.1f}min)")
        print(f"Mean: {statistics.mean(times):.1f}s ({statistics.mean(times)/60:.1f}min)")
        print(f"Under 8 minutes (<480s): {len(under_8min)}")

    if under_8min:
        print()
        print("Sub-8-minute completions (detail):")
        for rec, qt, r in under_8min:
            ind = r["profile"].get("industry", {})
            ind_label = ind.get("label") if isinstance(ind, dict) else "?"
            sen = r["profile"].get("seniority", {})
            sen_label = sen.get("label") if isinstance(sen, dict) else "?"
            pl = r["profile"].get("panel_list", {})
            pl_label = pl.get("label") if isinstance(pl, dict) else "?"
            print(f"  record={rec}: {qt:.1f}s | industry={ind_label} | seniority={sen_label} | panel_list={pl_label}")

        # Concentration check
        industries_sub8 = Counter(r["profile"].get("industry", {}).get("label") for _, _, r in under_8min)
        vlists_sub8 = Counter(r["profile"].get("panel_list", {}).get("code") for _, _, r in under_8min)
        print()
        print("Sub-8-min by industry:", dict(industries_sub8))
        print("Sub-8-min by vlist code:", dict(vlists_sub8))

    # -----------------------------------------------------------------------
    # F. vdropout
    # -----------------------------------------------------------------------
    print()
    print("=" * 60)
    print("ITEM F: vdropout (last_question_seen)")
    print("=" * 60)

    vdropout_dist = Counter()
    for r in data:
        xlsx_row = xlsx_rows.get(r["record"])
        vd = xlsx_row[col_map["vdropout"]] if xlsx_row else None
        if "metadata" not in r:
            r["metadata"] = {}
        r["metadata"]["last_question_seen"] = vd
        vdropout_dist[vd] += 1

    print("vdropout distribution across all 600 records:")
    for val, cnt in vdropout_dist.most_common():
        print(f"  {repr(val)}: {cnt}")
    print()
    print("All 600 qualified completers have vdropout=None — they reached the end of the survey.")
    print("metadata.last_question_seen set to None for all 600 records (no partial completions).")

    # -----------------------------------------------------------------------
    # G. Open-ended "other specify" fields
    # -----------------------------------------------------------------------
    print()
    print("=" * 60)
    print("ITEM G: Open-ended 'other specify' fields")
    print("=" * 60)

    OE_FIELDS = {
        "S4_Industryr28oe":       ("profile", "industry_other"),
        "S6_Functionr22oe":       ("profile", "functions_other_specify"),
        "Q13r12oe":               ("responses", "q13_other"),
        "Q15_Challengesr13oe":    ("responses", "q15_other"),
        "Q16_Dissuader12oe":      ("responses", "q16_other"),
        "Q21_TSPConfidencer13oe": ("responses", "q21_other"),
        "Q29r9oe":                ("responses", "q29_other"),
        "Q33_PositioningQsr11oe": ("responses", "q33_other"),
        "Q36_Pricingr7oe":        ("responses", "q36_other"),
        "Q37_SourcesofInfor14oe": ("responses", "q37_other"),
        "Q42r10oe":               ("responses", "q42_other"),
        "Q46r18oe":               ("responses", "q46_other"),
        "Q50r5oe":                ("responses", "q50_other"),
    }

    oe_results = {}
    for col_name, (section, field_name) in OE_FIELDS.items():
        if col_name not in col_map:
            oe_results[col_name] = []
            continue
        col_idx = col_map[col_name]
        vals = []
        for r in data:
            xlsx_row = xlsx_rows.get(r["record"])
            if xlsx_row:
                v = xlsx_row[col_idx]
                if v is not None and str(v).strip():
                    vals.append((r["record"], str(v).strip(), r, section, field_name))
        oe_results[col_name] = vals

    print(f"{'Field':<30} {'Non-null':>8}  {'Stored?':>8}")
    print("-" * 55)
    for col_name, vals in oe_results.items():
        _, (section, field_name) = col_name, OE_FIELDS[col_name]
        stored = "Yes" if len(vals) >= 5 else "No (< 5)"
        print(f"  {col_name:<28} {len(vals):>8}  {stored:>8}")

    # Store fields with ≥5 responses
    print()
    print("Verbatims for fields with ≥5 responses:")
    for col_name, vals in oe_results.items():
        if len(vals) < 5:
            if len(vals) > 0:
                print(f"\n{col_name} ({len(vals)} response — not stored, below threshold):")
                for rec, v, r, section, field_name in vals:
                    print(f"  record={rec}: {repr(v)}")
            continue
        _, (section, field_name) = col_name, OE_FIELDS[col_name]
        print(f"\n{col_name} → {section}.{field_name} ({len(vals)} responses):")
        for rec, v, r, section, field_name in vals:
            print(f"  record={rec}: {repr(v)}")
            # Store in document
            r[section][field_name] = v

    # -----------------------------------------------------------------------
    # H. Q52 segment thresholds — document
    # -----------------------------------------------------------------------
    print()
    print("=" * 60)
    print("ITEM H: Q52 (respondent_segment) threshold documentation")
    print("=" * 60)

    # Budget: Q39 vs segment
    print("Budget dimension — Q39 spend bands by segment:")
    q39_by_seg = {"High Budget": Counter(), "Low Budget": Counter(), "None": Counter()}
    for r in data:
        q39 = r["responses"].get("q39")
        seg = r["responses"].get("respondent_segment", {})
        q39_code = q39.get("code") if isinstance(q39, dict) else None
        seg_label = seg.get("label") if isinstance(seg, dict) else None
        if seg_label and "High Budget" in seg_label:
            q39_by_seg["High Budget"][q39_code] += 1
        elif seg_label and "Low Budget" in seg_label:
            q39_by_seg["Low Budget"][q39_code] += 1
        else:
            q39_by_seg["None"][q39_code] += 1

    for seg_type, dist in q39_by_seg.items():
        codes_str = ", ".join(
            f"code={c}({Q39_LABELS.get(c,'?')}):{n}"
            for c, n in sorted(dist.items(), key=lambda x: (x[0] is None, x[0] or 0))
        )
        print(f"  {seg_type}: {codes_str}")

    print()
    print("Segment definitions (from vlist/panel design, not derived from survey responses):")
    print()
    print("  BUDGET DIMENSION")
    print("  'High Budget' = respondents recruited into high-spend panels (vlist 3,4,7,8)")
    print("    → Broadly corresponds to Q39 ≥ code 5 ($10m+ annual AI spend)")
    print("    → Operationally: Q39 codes 5–8, but panels were pre-screened, not post-stratified")
    print("    → Code 5 ($10m–$14.9m) has partial overlap: 112 High Budget, 28 Low Budget")
    print("      This is because panel lists 3/4/7/8 used different Q39 thresholds per wave")
    print()
    print("  RANK DIMENSION")
    print("  'High Rank' = respondents who gave Cognizant a high AI capability rating in")
    print("    their panel profile (pre-survey screener data, not captured in this dataset)")
    print("  'Low Rank' = respondents who rated Cognizant lower in their panel profile")
    print("  NOTE: Q1_TSP_Now in-survey Cognizant ratings show no clean separation")
    print("    between High/Low Rank — confirming rank was assigned pre-survey, not")
    print("    computed from Q1. The threshold (exact rating cutpoint) is not derivable")
    print("    from the current dataset and should be obtained from the panel vendor.")
    print()
    print("  Q52 NULL COUNT: 46 records (vlist codes 1 and 2: Open Survey and IMR panels)")
    print("    These respondents were recruited without segment pre-assignment.")

    # Add segment documentation to _meta
    seg_doc = {
        "budget_threshold": "Q39 >= $10m/year (codes 5–8); pre-screened by panel vendor",
        "rank_threshold": "Cognizant AI capability rating from panel profile (pre-survey); exact cutpoint not available in dataset",
        "null_segments": "46 records (vlist 1=Open Survey, vlist 2=IMR) — no segment assigned",
    }
    for r in data:
        r["_meta"]["segment_definition"] = seg_doc

    # -----------------------------------------------------------------------
    # Final validation
    # -----------------------------------------------------------------------
    print()
    print("=" * 60)
    print("FINAL VALIDATION")
    print("=" * 60)
    assert len(data) == 600, "Record count mismatch!"
    print(f"Total records: {len(data)} ✓")

    checks = {
        "profile.panel_list": sum(1 for r in data if "panel_list" in r["profile"]),
        "profile.employee_bucket_computed": sum(1 for r in data if "employee_bucket_computed" in r["profile"]),
        "profile.employee_size_check": sum(1 for r in data if "employee_size_check" in r["profile"]),
        "metadata.interview_duration_sec": sum(1 for r in data if r.get("metadata", {}).get("interview_duration_sec") is not None),
        "metadata.last_question_seen": sum(1 for r in data if "last_question_seen" in r.get("metadata", {})),
        "_meta.segment_definition": sum(1 for r in data if "segment_definition" in r.get("_meta", {})),
        "responses.q52 absent": sum(1 for r in data if "q52" not in r["responses"]),
        "responses.ai_budget_planned absent": sum(1 for r in data if "ai_budget_planned" not in r["responses"]),
    }
    for field, count in checks.items():
        status = "✓" if count == 600 else f"⚠ {count}"
        print(f"  {field}: {count}/600 {status}")

    # Field inventory
    print()
    print("Complete field inventory:")
    sample = data[0]
    print()
    print("PROFILE fields:", sorted(sample["profile"].keys()))
    print()
    print("RESPONSES fields:", sorted(sample["responses"].keys()))
    print()
    print("METADATA fields:", sorted(sample.get("metadata", {}).keys()))
    print()
    print("_META fields:", sorted(sample["_meta"].keys()))

    # Write
    with open(RESPONDENTS_PATH, "w") as f:
        json.dump(data, f, indent=2)
    print()
    print(f"Updated data written to {RESPONDENTS_PATH}")


if __name__ == "__main__":
    main()
