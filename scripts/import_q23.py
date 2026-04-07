#!/usr/bin/env python3
"""
Import Q23 (Overall AI perception, -3 to +3) from xlsx into responses.q23.
Maps by record number. Only stores entries where value is non-null (shown brands).
"""
import json
import openpyxl

RESPONDENTS_PATH = "data/respondents.json"
XLSX_PATH = "/Users/972010/Downloads/Cognizant_Raw_Data.xlsx"

BRAND_IDX_TO_NAME = {
    1: "Cognizant",
    2: "Accenture",
    3: "IBM Consulting",
    4: "Infosys",
    5: "Capgemini",
    6: "Wipro",
    7: "Tata Consultancy Services (TCS)",
    8: "EY",
    9: "HCL Technologies",
    10: "Deloitte",
    11: "McKinsey & Company",
    12: "Google (Cloud & Gemini)",
    13: "DXC Technology",
    14: "ServiceNow",
    15: "Microsoft (Azure & Copilot)",
    16: "Amazon Web Services (AWS)",
}

print("Loading xlsx...")
wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
ws = wb["A1"]
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
rec_col = headers.index("record")
q23_cols = [(i, int(h.replace("Q23_Perceptionr", "")))
            for i, h in enumerate(headers) if h and str(h).startswith("Q23_Perceptionr")]
q23_cols.sort(key=lambda x: x[1])  # sort by brand_idx

# Build map: record → list of {brand_idx, brand, value}
xlsx_q23 = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    rec = row[rec_col]
    if rec is None:
        continue
    entries = []
    for col_i, brand_idx in q23_cols:
        val = row[col_i]
        if val is not None:
            entries.append({
                "brand_idx": brand_idx,
                "brand": BRAND_IDX_TO_NAME[brand_idx],
                "value": int(val),
                "question_code": f"Q23_Perceptionr{brand_idx}",
            })
    xlsx_q23[int(rec)] = entries
wb.close()
print(f"Loaded Q23 data for {len(xlsx_q23)} records from xlsx.")

# Load JSON
with open(RESPONDENTS_PATH) as f:
    data = json.load(f)
assert len(data) == 600

updated = 0
total_entries = 0
for r in data:
    rec = r["record"]
    entries = xlsx_q23.get(rec, [])
    r["responses"]["q23"] = entries
    total_entries += len(entries)
    if entries:
        updated += 1

print(f"Records with ≥1 Q23 value: {updated}/600")
print(f"Total brand-level Q23 entries: {total_entries}")
print(f"Avg entries per non-empty record: {total_entries/updated:.1f}" if updated else "")

with open(RESPONDENTS_PATH, "w") as f:
    json.dump(data, f, indent=2)
print(f"Written to {RESPONDENTS_PATH}")
